# -*- coding:utf-8 -*-#

from __future__ import division

import sys
reload(sys)
sys.setdefaultencoding('utf8')

import json,pickle,random,Levenshtein,re
import tornado.web
import db_sql as db
from openpyxl import load_workbook
import os

# import lh_search.lh_search as lh_search
# from zhenduan_classifier.lph_zhenduan import lph_nb_rules_classifier

# from knowledge_graphs import lph_knowledge_reasoning
from info_parse import run_parse


class IndexHandler(tornado.web.RequestHandler):
    def get(self):
        print '----------Index GET----------'
        if len(self.request.query) == 0:
            path = 'static/lung_project/files/5000processed.xlsx'
        else:
            print self.request.query
            path = 'static/lung_project/files/' + self.request.query.split('=')[1]
        workbook = load_workbook(path)
        sheet = workbook.active

        is_m = 0
        not_m = 0
        other = 0
        left_not = 0
        left_up_not = 0
        left_down_not = 0
        right_not = 0
        right_up_not = 0
        right_mid_not = 0
        right_down_not = 0
        double_not = 0
        left = 0
        left_up = 0
        left_down = 0
        right = 0
        right_up = 0
        right_mid = 0
        right_down = 0
        double = 0

        l_multy_yes = []
        l_multy_not = []

        for i in range(sheet.max_row - 2):
            if str(sheet.cell(row=i + 3, column=4).value).find('0') != -1:
                other += 1
            if str(sheet.cell(row=i + 3, column=4).value).find('1') != -1:
                is_m += 1
                if sheet.cell(row=i + 3, column=14).value is not None:
                    left += 1
                if sheet.cell(row=i + 3, column=15).value is not None:
                    left_up += 1
                if sheet.cell(row=i + 3, column=16).value is not None:
                    left_down += 1
                if sheet.cell(row=i + 3, column=17).value is not None:
                    right += 1
                if sheet.cell(row=i + 3, column=18).value is not None:
                    right_up += 1
                if sheet.cell(row=i + 3, column=19).value is not None:
                    right_mid += 1
                if sheet.cell(row=i + 3, column=20).value is not None:
                    right_down += 1
                if sheet.cell(row=i + 3, column=21).value is not None:
                    double += 1
            if str(sheet.cell(row=i + 3, column=4).value).find('2') != -1:
                not_m += 1
                if sheet.cell(row=i + 3, column=5).value is not None:
                    left_not += 1
                if sheet.cell(row=i + 3, column=6).value is not None:
                    left_up_not += 1
                if sheet.cell(row=i + 3, column=7).value is not None:
                    left_down_not += 1
                if sheet.cell(row=i + 3, column=8).value is not None:
                    right_not += 1
                if sheet.cell(row=i + 3, column=9).value is not None:
                    right_up_not += 1
                if sheet.cell(row=i + 3, column=10).value is not None:
                    right_mid_not += 1
                if sheet.cell(row=i + 3, column=11).value is not None:
                    right_down_not += 1
                if sheet.cell(row=i + 3, column=12).value is not None:
                    double_not += 1

        threshold = 0.065   # 判定多发阈值

        if left/is_m > threshold:
            l_multy_yes.append(0)
        if left_up/is_m > threshold:
            l_multy_yes.append(1)
        if left_down/is_m > threshold:
            l_multy_yes.append(2)
        if right/is_m > threshold:
            l_multy_yes.append(3)
        if right_up/is_m > threshold:
            l_multy_yes.append(4)
        if right_mid/is_m > threshold:
            l_multy_yes.append(5)
        if right_down/is_m > threshold:
            l_multy_yes.append(6)
        if double/is_m > threshold:
            l_multy_yes.append(7)
        if left_not/not_m > threshold:
            l_multy_not.append(0)
        if left_up_not/not_m > threshold:
            l_multy_not.append(1)
        if left_down_not/not_m > threshold:
            l_multy_not.append(2)
        if right_not/not_m > threshold:
            l_multy_not.append(3)
        if right_up_not/not_m > threshold:
            l_multy_not.append(4)
        if right_mid_not/not_m > threshold:
            l_multy_not.append(5)
        if right_down_not/not_m > threshold:
            l_multy_not.append(6)
        if double_not/not_m > threshold:
            l_multy_not.append(7)

        if len(self.request.query) == 0:    # 用户点击侧栏“胸部CT处理统计”按钮
            filename = ''
        else:   # 上传文件跳转到index页面
            filename = self.request.query.split('=')[1]
        self.render("doctor_home.html", all_cases={}, current_doctor={'name': '万医生'}, source=0, filename=filename,
                is_m=is_m, not_m=not_m, other=other, left=left, left_up=left_up, left_down=left_down, right=right,
                right_up=right_up, right_mid=right_mid, right_down=right_down, double=double,
                left_not=left_not, left_up_not=left_up_not, left_down_not=left_down_not, right_not=right_not,
                right_up_not=right_up_not, right_mid_not=right_mid_not, right_down_not=right_down_not, double_not=double_not,
                l_multy_yes=l_multy_yes, l_multy_not=l_multy_not)

    def post(self):
        print '----------Index POST----------'
        if len(self.get_argument('filename')) == 0:
            path = 'static/lung_project/files/5000processed.xlsx'
        else:
            path = 'static/lung_project/files/' + self.get_argument('filename')
        workbook = load_workbook(path)
        sheet = workbook.active

        d_size = dict()

        info = self.get_argument('info')
        if info == '磨玻璃':
            for i in range(sheet.max_row - 2):
                for j in range(14, 22):
                    if sheet.cell(row=i+3, column=j).value is not None and str(sheet.cell(row=i+3, column=j).value) != 'A' and str(sheet.cell(row=i+3, column=j).value) != 'M':
                        size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=j).value)).group())
                        if size not in d_size.keys():
                            d_size[size] = 1
                        else:
                            d_size[size] += 1
        elif info == '非磨玻璃':
            for i in range(sheet.max_row - 2):
                for j in range(5, 13):
                    if sheet.cell(row=i+3, column=j).value is not None and str(sheet.cell(row=i+3, column=j).value) != 'A' and str(sheet.cell(row=i+3, column=j).value) != 'M':
                        size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=j).value)).group())
                        if size not in d_size.keys():
                            d_size[size] = 1
                        else:
                            d_size[size] += 1
        elif info == '左肺非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=5).value is not None and str(sheet.cell(row=i+3, column=5).value) != 'A' and str(sheet.cell(row=i+3, column=5).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=5).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '左上非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=6).value is not None and str(sheet.cell(row=i+3, column=6).value) != 'A' and str(sheet.cell(row=i+3, column=6).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=6).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '左下非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=7).value is not None and str(sheet.cell(row=i+3, column=7).value) != 'A' and str(sheet.cell(row=i+3, column=7).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=7).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右肺非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=8).value is not None and str(sheet.cell(row=i+3, column=8).value) != 'A' and str(sheet.cell(row=i+3, column=8).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=8).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右上非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=9).value is not None and str(sheet.cell(row=i+3, column=9).value) != 'A' and str(sheet.cell(row=i+3, column=9).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=9).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右中非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=10).value is not None and str(sheet.cell(row=i+3, column=10).value) != 'A' and str(sheet.cell(row=i+3, column=10).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=10).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右下非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=11).value is not None and str(sheet.cell(row=i+3, column=11).value) != 'A' and str(sheet.cell(row=i+3, column=11).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=11).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '两肺非':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=12).value is not None and str(sheet.cell(row=i+3, column=12).value) != 'A' and str(sheet.cell(row=i+3, column=12).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=12).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '左肺':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=14).value is not None and str(sheet.cell(row=i+3, column=14).value) != 'A' and str(sheet.cell(row=i+3, column=14).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=14).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '左上':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=15).value is not None and str(sheet.cell(row=i+3, column=15).value) != 'A' and str(sheet.cell(row=i+3, column=15).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=15).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '左下':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=16).value is not None and str(sheet.cell(row=i+3, column=16).value) != 'A' and str(sheet.cell(row=i+3, column=16).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=16).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右肺':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=17).value is not None and str(sheet.cell(row=i+3, column=17).value) != 'A' and str(sheet.cell(row=i+3, column=17).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=17).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右上':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=18).value is not None and str(sheet.cell(row=i+3, column=18).value) != 'A' and str(sheet.cell(row=i+3, column=18).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=18).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右中':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=19).value is not None and str(sheet.cell(row=i+3, column=19).value) != 'A' and str(sheet.cell(row=i+3, column=19).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=19).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '右下':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=20).value is not None and str(sheet.cell(row=i+3, column=20).value) != 'A' and str(sheet.cell(row=i+3, column=20).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=20).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1
        elif info == '两肺':
            for i in range(sheet.max_row - 2):
                if sheet.cell(row=i+3, column=21).value is not None and str(sheet.cell(row=i+3, column=21).value) != 'A' and str(sheet.cell(row=i+3, column=21).value) != 'M':
                    size = float(re.search('\d+\.?\d*', str(sheet.cell(row=i + 3, column=21).value)).group())
                    if size not in d_size.keys():
                        d_size[size] = 1
                    else:
                        d_size[size] += 1

        # l_id = []
        # l_desc = []
        # l_diag = []
        #
        # info = self.get_argument('info')
        # if info == '磨玻璃':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '非磨玻璃':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '其他':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 0:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '左肺非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and not sheet.cell(row=i+3, column=5).value:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '左上非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and not sheet.cell(row=i+3, column=6).value:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '左下非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and sheet.cell(row=i+3, column=7).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右肺非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and sheet.cell(row=i+3, column=8).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右上非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and sheet.cell(row=i+3, column=9).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右中非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and sheet.cell(row=i+3, column=10).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右下非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and sheet.cell(row=i+3, column=11).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '两肺非':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 2 and sheet.cell(row=i+3, column=12).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '左肺':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=14).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '左上':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=15).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '左下':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=16).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右肺':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=17).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右上':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=18).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右中':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=19).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '右下':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=20).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        # elif info == '两肺':
        #     for i in range(sheet.max_row - 2):
        #         if sheet.cell(row=i+3, column=4).value == 1 and sheet.cell(row=i+3, column=21).value is not None:
        #             l_id.append(sheet.cell(row=i + 3, column=1).value)
        #             l_desc.append(sheet.cell(row=i + 3, column=2).value)
        #             l_diag.append(sheet.cell(row=i + 3, column=3).value)
        #
        # all_cases = {
        #     'l_id': l_id,
        #     'l_desc': l_desc,
        #     'l_diag': l_diag
        # }  # 封装数据
        #
        # f = open('static/lung_project/files/tableData.json', "w")
        # for i in range(len(all_cases['l_id'])):
        #     json.dump({"编号": all_cases['l_id'][i], "胸部CT所见": all_cases['l_desc'][i], "胸部CT诊断": all_cases['l_diag'][i]}, f)

        all_cases = {
            'd_size': d_size
        }  # 封装数据

        self.write(json.dumps(all_cases))


class UploadHandler(tornado.web.RequestHandler):
    def get(self):
        print '--------Upload GET---------'
        filetype = 1
        if len(self.request.query) != 0:    # 文件不是excel
            filetype = 0
            print 'invalid file type'
        self.render("upload.html", all_cases={}, current_doctor={'name': '万医生'}, source=0, filetype=filetype)

    def post(self):
        print '--------Upload POST你是怎么找到我的？---------'


class UploadApiHandler(tornado.web.RequestHandler):
    def get(self):
        print '--------UploadApi GET---------'
    def post(self):
        print '--------UploadApi POST---------'
        ret = {'result': 'OK'}
        upload_path = os.path.join(os.path.dirname(__file__), r"static\lung_project\files")  # 文件存储路径
        file_metas = self.request.files.get('file', None)  # [{'body':'', 'content_type':'', 'filename':'xxx.xxx'}]

        if not file_metas:
            ret['result'] = 'Invalid Args'
            return ret

        file = file_metas[0]        # {'body':'', 'content_type':'', 'filename':'xxx.xxx'}
        filename = file['filename']
        if filename.find('xlsx') != -1:
            filename = str(hash(filename.split('.')[0])) + '.' + str(filename.split('.')[1])
            file_path = os.path.join(upload_path, filename)
            with open(file_path, 'wb') as f:
                f.write(file['body'])

            self.write(json.dumps(ret))

            url = '/lung/index?' + 'path=' + filename
            self.redirect(url)
        else:
            url = '/lung/upload?type=invalid'
            self.redirect(url)


# class SearchHandler(tornado.web.RequestHandler):
#     xitong = pickle.load(open("./data/backend/xitong_new.pkl", "r"))
#     remen = pickle.load(open("./data/backend/remen_new.pkl", "r"))
#     source = ''
#
#     def get(self):
#         query_arguments = self.request.arguments
#         # print query_arguments
#         if len(query_arguments) == 0:
#             self.render("xy_search_home.html", query=None, search_results=None, xitong=self.xitong, remen=self.remen,source=self.source)
#         else:
#             query = self.get_argument("query")
#             cur_page = int(query_arguments.get("page", [1])[0])
#             search_results = lh_search.run_search(query)
#             page_size = 10
#             pages = len(search_results) / page_size + 1
#             # tishi = lph_knowledge_reasoning.get_zhengzhuang_tishi(query)
#             tishi_res = lph_nb_rules_classifier.lph_wenda_zhenduan_for_search(query)
#             tishi = tishi_res.split(u'，')[0:8]
#             self.render("xy_search_home.html", cur_page=cur_page, page_size=page_size, pages=pages,
#                         query=query, search_results=search_results, tishi=tishi,
#                         xitong=self.xitong, remen=self.remen,source=self.source)
#     # xitong = pickle.load(open("./data/backend/xitong_new.pkl", "r"))
#     # remen = pickle.load(open("./data/backend/remen_new.pkl", "r"))
#     #
#     # def get(self):
#     #     query_arguments = self.request.arguments
#     #     # print query_arguments
#     #     if len(query_arguments) == 0:
#     #         self.render("xy_search_home.html", query=None, search_results=None, xitong=self.xitong, remen=self.remen,source = 'nwx')
#     #     else:
#     #         query = self.get_argument("query")
#     #         cur_page = int(query_arguments.get("page", [1])[0])
#     #         # search_results = lh_search.run_search(query)
#     #         search_results = ['one','two']
#     #         page_size = 2
#     #         # pages = len(search_results) / page_size + 1
#     #         pages = 1
#     #         # tishi = lph_knowledge_reasoning.get_zhengzhuang_tishi(query)
#     #         # tishi_res = lph_nb_rules_classifier.lph_wenda_zhenduan_for_search(query)
#     #         # tishi = tishi_res.split(u'，')[0:8]
#     #         tishi = ['tip1','tip2']
#     #         self.render("xy_search_home.html", cur_page=cur_page, page_size=page_size, pages=pages,
#     #                     query=None, search_results=search_results, tishi=tishi,
#     #                     xitong=self.xitong, remen=self.remen,source="nwx")
#     # def post(self):
#     #     source = "nwx"
#     #     query_arguments = self.request.arguments
#     #     query = self.get_argument("query")
#     #     cur_page = int(query_arguments.get("page", [1])[0])
#     #     # search_results = lh_search.run_search(query)
#     #     search_results = ['one', 'two']
#     #     page_size = 2
#     #     # pages = len(search_results) / page_size + 1
#     #     pages = 1
#     #     # tishi = lph_knowledge_reasoning.get_zhengzhuang_tishi(query)
#     #     # tishi_res = lph_nb_rules_classifier.lph_wenda_zhenduan_for_search(query)
#     #     # tishi = tishi_res.split(u'，')[0:8]
#     #     tishi = ['tip1', 'tip2']
#     #     self.render("xy_search_home.html", cur_page=cur_page, page_size=page_size, pages=pages,
#     #                 query=query, search_results=search_results, tishi=tishi,
#     #                 xitong=self.xitong, remen=self.remen, source="nwx")
#     #     self.render("xy_search_home.html",source=source,query=None)
#
#
# class LoginHandler(tornado.web.RequestHandler):
#     """
#     登录逻辑控制器
#     """
#     with open("./data/backend/all_validate_images.pkl", "r") as f:
#         all_validata_images = pickle.load(f)
#
#     def get(self):
#         print "login"
#         image_info = random.choice(self.all_validata_images)
#         self.render("login.html", img_info=image_info)
#
#     def post(self):
#         user_type = self.get_argument("user_type")
#         user_phone = self.get_argument("user_phone")
#         user_password = self.get_argument("user_password")
#         user_id, user_name, is_existed = db.search_user(user_phone, user_password, int(user_type))
#         res_json = {}
#         if is_existed == 1:
#             res_json["status"] = "success"
#             res_json["user_type"] = int(user_type)
#         else:
#             res_json["status"] = "fail"
#         self.set_secure_cookie("user_id", str(user_id))
#         print user_name,type(user_name)
#         self.set_secure_cookie("user_name", (user_name))
#         self.set_secure_cookie("user_type", str(user_type))
#         self.write(json.dumps(res_json))
#
#
# class GraphHandler(tornado.web.RequestHandler):
#     def get(self):
#         p_id = self.get_secure_cookie("user_id")
#         p_type = self.get_secure_cookie("user_type")
#         print p_type,p_id
#         if p_type == "1" and p_id is not None:
#             doctor_dict = {
#                 u'doctorID': int(p_id),
#             }
#             res_doctor = db.query_doctor(doctor_dict)[0]
#             self.render("knowledge_graph_home.html",source ="",current_doctor = {u'name':"江院长"})
#         else:
#             self.redirect("/shanghai/login")
#
#     def post(self):
#         self.render("knowledge_graph_home.html",source='')
#
#
# class GetAllGraphHandler(tornado.web.RequestHandler):
#     """
#     读取所有疾病的知识图谱信息
#     """
#
#     all_examine_disease = pickle.load(open(u"knowledge_graphs/data/cx_all_examine_disease.pkl", 'r'))
#
#     def post(self):
#         page = int(self.get_argument("page")) - 1
#         # start = page * 2
#         # end = start + 2
#         all_graph = lph_knowledge_reasoning.read_all()
#         result = {
#             u"all_graph": all_graph.values()[page:page+1],
#             u"all_examine_disease": self.all_examine_disease
#         }
#         self.write(json.dumps(result))
#
#
# class SearchGraphHandler(tornado.web.RequestHandler):
#     """
#     yisheng就诊页面，问答的形式
#     """
#     # hmm_all_graph_body = pickle.load(open(u"knowledge_graphs/data/erke_book_graph.pkl", "r"))
#     hmm_all_graph_body = pickle.load(open(u"knowledge_graphs/data/erke_book_graph_refine_with_yuanwen.pkl", "r"))
#     # refine_book_graph = pickle.load(open(u"knowledge_graphs/data/erke_hmm_graph.pkl", "r"))
#     all_graph_body = pickle.load(open(u"knowledge_graphs/data/cx_all_graph.pkl", "r"))
#     skin_graph_body = pickle.load(open(u"knowledge_graphs/data/skin_baike_graph.pkl", "r"))
#     biyan_graph_body = pickle.load(open(u"knowledge_graphs/data/biyan_baike_graph.pkl", "r"))
#     book_skin_graph_body = pickle.load(open(u"knowledge_graphs/data/skin_book_graph.pkl", "r"))
#     symptom_xuewei_all_graph_body = pickle.load(open(u"knowledge_graphs/data/symptom_xuewei_graphs.pkl", "r"))
#     def post(self):
#         id_name = self.get_argument("id_name")
#         graph_name = self.get_argument("graph_name")
#         # print "graph_name",json.dumps(graph_name,ensure_ascii=False,encoding=u'utf-8')
#         result_graph = self.get_graph(id_name, graph_name)
#         self.write(json.dumps(result_graph))
#
#     def get_graph(self, id_name, graph_name):
#         # path_dir = u""
#         # if id_name == u"hmm_all_graph_body":
#         #     path_dir = u"hmm_pkl_graphs"
#         # elif id_name == u"all_graph_body":
#         #     path_dir = u"cx_pkl_graphs"
#         # elif id_name == u"skin_graph_body":
#         #     path_dir = u"skin_pkl_graphs"
#         # elif id_name == u"biyan_graph_body":
#         #     path_dir = u"biyan_pkl_graphs"
#         # elif id_name == u"book_skin_graph_body":
#         #     path_dir = u"book_skin_pkl_graphs"
#         # elif id_name == u"symptom_xuewei_all_graph_body":
#         #     path_dir = u"symptom_xuewei_graphs"
#
#         if id_name == u"hmm_all_graph_body":
#             refine_book_graph = self.hmm_all_graph_body
#         elif id_name == u"all_graph_body":
#             # refine_book_graph = pickle.load(open(u"knowledge_graphs/data/erke_hmm_graph.pkl", "r"))
#             refine_book_graph = self.all_graph_body
#         elif id_name == u"skin_graph_body":
#             refine_book_graph = self.skin_graph_body
#         elif id_name == u"biyan_graph_body":
#             refine_book_graph = self.biyan_graph_body
#         elif id_name == u"book_skin_graph_body":
#             refine_book_graph = self.book_skin_graph_body
#         elif id_name == u"symptom_xuewei_all_graph_body":
#             refine_book_graph = self.symptom_xuewei_all_graph_body
#         # result = self.get_graph_name(id_name, graph_name)[0]
#         # print 'graph_name',json.dumps(result,ensure_ascii=False,encoding=u'utf-8')
#         graph_keys = refine_book_graph.keys()
#         res = {}
#         if graph_name in graph_keys:
#             return [1,refine_book_graph.get(graph_name)]
#         for key in graph_keys:
#             if graph_name.strip(u'急性') in key or graph_name.strip(u'先天性') in key or graph_name.strip(u'新生儿') in key:
#                 return [1, refine_book_graph.get(key)]
#             if key.strip(u'急性') in graph_name or key.strip(u'先天性') in graph_name or key.strip(u'新生儿') in graph_name:
#                 return [1, refine_book_graph.get(key)]
#             res[key] = Levenshtein.ratio(key, graph_name)
#         res_list = sorted(res.items(), key=lambda item: item[1], reverse=True)
#         result = res_list[0][0]
#         try:
#             graph = refine_book_graph.get(result)
#         except Exception:
#             return [-1, {}]
#         return [1, graph]
#
#     # def get_graph_name(self, id_name, graph_name):
#     #     if id_name == u"hmm_all_graph_body":
#     #         refine_book_graph = pickle.load(open(u"knowledge_graphs/data/erke_book_graph.pkl", "r"))
#     #     elif id_name == u"all_graph_body":
#     #         # refine_book_graph = pickle.load(open(u"knowledge_graphs/data/erke_hmm_graph.pkl", "r"))
#     #         refine_book_graph = pickle.load(open(u"knowledge_graphs/data/cx_all_graph.pkl", "r"))
#     #     elif id_name == u"skin_graph_body":
#     #         refine_book_graph = pickle.load(open(u"knowledge_graphs/data/skin_baike_graph.pkl", "r"))
#     #     elif id_name == u"biyan_graph_body":
#     #         refine_book_graph = pickle.load(open(u"knowledge_graphs/data/biyan_baike_graph.pkl", "r"))
#     #     elif id_name == u"book_skin_graph_body":
#     #         refine_book_graph = pickle.load(open(u"knowledge_graphs/data/skin_book_graph.pkl", "r"))
#     #     elif id_name == u"symptom_xuewei_all_graph_body":
#     #         refine_book_graph = pickle.load(open(u"knowledge_graphs/data/symptom_xuewei_graphs.pkl","r"))
#     #     graph_keys = refine_book_graph.keys()
#     #     res = {}
#     #     if graph_name in graph_keys:
#     #         return graph_name
#     #     for key in graph_keys:
#     #         if graph_name.strip(u'急性') in key or graph_name.strip(u'先天性') in key or graph_name.strip(u'新生儿') in key:
#     #             return [key]
#     #         if key.strip(u'急性') in graph_name or key.strip(u'先天性') in graph_name or key.strip(u'新生儿') in graph_name:
#     #             return [key]
#     #         res[key] = Levenshtein.ratio(key, graph_name)
#     #     res_list = sorted(res.items(), key=lambda item: item[1], reverse=True)
#     #     result = []
#     #     for i in res_list[:5]:
#     #         result.append(i[0])
#     #     return result
#
