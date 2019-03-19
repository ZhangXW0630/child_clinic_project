# -*- coding:utf-8 -*-#

import sys
reload(sys)
sys.setdefaultencoding('utf8')

import re
from openpyxl import load_workbook


def process_excel(path):
    workbook = load_workbook(path)
    sheet = workbook.active  # 获取当前活跃的Worksheet
    # 读
    # print(sheet.cell(row=2, column=2).value)    # row从1开始
    # print(sheet['B3'].value)
    # 写
    # sheet['A15'] = '47'
    # sheet.cell(row=20, column=1).value = 2

    for i in range(sheet.max_row - 2):
        # ------------------------遍历每一行----------------------------
        # 结节分类
        flag = ''
        wrongPosition = True  # 关键词不是肺部的（错误位置）
        descp_str = str(sheet.cell(row=i + 3, column=2).value)
        diag_str = str(sheet.cell(row=i + 3, column=3).value)
        print i + 1, '---------------------->',

        # 匹配磨玻璃关键词
        kws = re.findall("磨玻璃", diag_str)
        if len(kws) != 0:
            print '磨玻璃'
            flag = '1'

        # 匹配非磨玻璃关键词
        if len(re.findall('密度稍高影|团块影|斑结影|斑结灶|团片状影|小斑结节影|(?<!磨玻璃)结节', diag_str)) != 0:
            sentences = re.split('，|。|；|\?|？', diag_str)  # 将CT诊断拆成单句
            for j, sentence in enumerate(sentences):
                # 检测结节是否在肺部
                if re.search('密度稍高影|团块影|斑结影|斑结灶|团片状影|小斑结节影|(?<!磨玻璃)结节', sentence) and re.search('肺', sentence):
                    wrongPosition = False  # 结节位置在肺部
                    # 检测是否“考虑炎性、陈旧病变”
                    if j != len(sentences) - 1 and re.search('考虑.*(?:炎性|陈旧)', sentences[j + 1]):  # 搜索下一句是否考虑炎性or陈旧
                        print('考虑陈旧or炎性。')
                    else:
                        print('非磨玻璃')
                        if flag == '':
                            flag = '2'
                        else:
                            flag += '、2'
                        break
            if wrongPosition:
                print('存在假关键词。')

        # 既不是磨玻璃，也不是非磨玻璃
        if flag == '':
            flag = '0'
            print('其它')

        # excel写入类别
        sheet.cell(row=i + 15, column=4).value = flag

        # 匹配位置、大小信息
        sentences = re.split('，|。|；|\?|？', descp_str)  # 诊断描述拆成单句
        left = False
        right = False
        double = False
        up = False
        mid = False
        down = False
        col = []
        if flag != '0':
            for j, sentence in enumerate(sentences):
                # ①非磨玻璃
                if re.search('密度稍高影|团块影|斑结影|斑结灶|团片状影|小斑结节影|斑节影|(?<!磨玻璃)结节', sentence) and not re.search('磨玻璃',
                                                                                                        sentence):  # No.11磨玻璃样结节
                    # 一句可能有多个位置，如第4条
                    positions = re.findall(
                        '(?:左|右|两)(?:上|中|下)?(?:肺)?(?:上叶|中叶|下叶|上、下叶|.{0,9}舌|.?后|前|背|.{0,2}基底|尖|.?侧)?(?:段)?', sentence)
                    if len(positions) != 0:  # 判断结节是否在肺部
                        if j != len(sentences) - 1 and re.search('考虑.*(?:炎性|陈旧)', sentences[j + 1]):  # 搜索下一句是否考虑炎性or陈旧
                            continue
                        else:  # 有结节位置且信息正确
                            print '第%d句：' % j, 'positions =', str(positions).decode('string_escape')
                            for k in positions:
                                if k.find('左肺') != -1:
                                    left = True
                                if k.find('右肺') != -1:
                                    right = True
                                if k.find('两肺') != -1:
                                    double = True
                                if re.search('上叶', k) or re.search('前段|尖', k):
                                    up = True
                                if re.search('尖段|(?<!尖)后段', k) or re.search('右上肺', k):
                                    right = True
                                    up = True
                                if re.search('尖后段|舌段', k) or re.search('左上肺', k):
                                    left = True
                                    up = True
                                if re.search('中叶', k):
                                    mid = True
                                if re.search('侧段', k) or re.search('右中肺', k):
                                    right = True
                                    mid = True
                                if re.search('下叶', k) or re.search('背段', k) or re.search('[外后]基底段', k):
                                    down = True
                                if re.search('前基底段|(?<!前)内基底段', k) or re.search('右下肺', k):
                                    right = True
                                    down = True
                                if re.search('前内基底段', k) or re.search('左下肺', k):
                                    left = True
                                    down = True
                                if left:
                                    if up:
                                        col.append(6)
                                    elif down:
                                        col.append(7)
                                    else:
                                        col.append(5)
                                elif right:
                                    if up:
                                        col.append(9)
                                    elif mid:
                                        col.append(10)
                                    elif down:
                                        col.append(11)
                                    else:
                                        col.append(8)
                                elif double:
                                    col.append(12)
                                left = False  # 标志位重置False
                                right = False
                                double = False
                                up = False
                                mid = False
                                down = False

                            # ------确定多发-------
                            isMulti = False  # 设置这个标志位默认前提条件：一句话中多发对应所有位置
                            if len(col) == 1 and re.search('\d[cm]*、\d[cm]+', sentence):  # 一个位置多个大小
                                sheet.cell(row=i + 15, column=col[0]).value = 'M'
                            if re.search('多发|[多数两三四]枚|少许', sentence):  # 存在多发、少许字样
                                for k in col:
                                    sheet.cell(row=i + 15, column=k).value = 'M'

                            # ------确定大小------
                            isSized = False  # 是否有大小信息
                            # ① 多个大小，顿号分隔(3mm、5mm)
                            if re.search('\d+\.?\d*[cm]*(?:、\d+\.?\d*[cm]*)+', sentence):
                                # （1）只有一个位置
                                if len(col) == 1:
                                    l_sizes = re.search('\d+\.?\d*[cm]*(?:、\d+\.?\d*[cm]*)+', sentence).group().split(
                                        '、')
                                    for n, size in enumerate(l_sizes):
                                        l_sizes[n] = float(re.search('[^mc]+', size).group())
                                    size_str = str(max(l_sizes))  # 最大值
                                    if re.search('mm', sentence):
                                        result = size_str
                                    else:  # cm
                                        if re.search('\.', size_str) and len(size_str.split('.')[1]) > 1:
                                            result = str(float(size_str) * 10)
                                        else:
                                            result = str(int(float(size_str) * 10))
                                # （2）多个位置
                                if len(positions):
                                    pass


                            # ② 单个大小的情况
                            reg = '\d*\.?\d*(?:-|\*)?\d*\.?\d*(?:mm|cm)'
                            size_str = re.search(reg, sentence)
                            if not size_str:
                                if j != len(sentences) - 1:  # 大小在下一句
                                    size_str = re.search(reg, sentences[j + 1])
                                    if not size_str:  # 没有大小信息填A
                                        for k in col:
                                            if not sheet.cell(row=i + 15, column=k).value:
                                                sheet.cell(row=i + 15, column=k).value = 'A'
                                    else:
                                        isSized = True
                            else:
                                isSized = True
                            if isSized:
                                size_str = size_str.group()
                                if re.search('-', size_str):  # 有-则取-后面的数作为大小
                                    result = re.search('(?<=-)\d*\.?\d*', size_str).group()
                                    if re.search('cm', size_str):
                                        if re.search('\.', result) and len(result.split('.')[1]) > 1:
                                            result = str(float(result) * 10)
                                        else:
                                            result = str(int(float(result) * 10))
                                elif re.search('\*', size_str):
                                    if re.search('mm', size_str):
                                        result = re.search('[^mc]+', size_str).group()
                                    else:  # cm
                                        size_str = re.search('[^mc]+', size_str).group()
                                        if re.search('\.', size_str.split('*')[0]) and len(
                                                size_str.split('*')[0].split('.')[1]) > 1:
                                            first = float(size_str.split('*')[0]) * 10
                                        else:
                                            first = int(float(size_str.split('*')[0]) * 10)
                                        if re.search('\.', size_str.split('*')[1]) and len(
                                                size_str.split('*')[1].split('.')[1]) > 1:
                                            second = float(size_str.split('*')[1]) * 10
                                        else:
                                            second = int(float(size_str.split('*')[1]) * 10)
                                        result = str(first) + '*' + str(second)
                                else:  # 单独数字
                                    if re.search('mm', size_str):
                                        result = re.search('[^mc]+', size_str).group()
                                    else:
                                        if re.search('\.', re.search('[^mc]+', size_str).group()) and len(
                                                re.search('[^mc]+', size_str).group().split('.')[1]) > 1:
                                            result = str(float(re.search('[^mc]+', size_str).group()) * 10)
                                        else:
                                            result = str(int(float(re.search('[^mc]+', size_str).group()) * 10))
                                for k in col:
                                    if sheet.cell(row=i + 15, column=k).value == 'M':
                                        sheet.cell(row=i + 15, column=k).value += result
                                    else:
                                        sheet.cell(row=i + 15, column=k).value = result
                            col = []  # 清空位置列表

                # ②磨玻璃
                if re.search('磨玻璃', sentence):
                    # 一句可能有多个位置，如第4条
                    positions = re.findall(
                        '(?:左|右|两)(?:上|中|下)?(?:肺)?(?:上叶|中叶|下叶|上、下叶|.{0,9}舌|.?后|前|背|.{0,2}基底|尖|.?侧)?(?:段)?', sentence)
                    for k in positions:
                        print '第%d句：' % j, 'positions =', positions
                        if k.find('左肺') != -1:
                            left = True
                        if k.find('右肺') != -1:
                            right = True
                        if k.find('两肺') != -1:
                            double = True
                        if re.search('上叶', k) or re.search('前段|尖', k):
                            up = True
                        if re.search('尖段|(?<!尖)后段', k):
                            right = True
                            up = True
                        if re.search('尖后段|舌段', k):
                            left = True
                            up = True
                        if re.search('中叶', k):
                            mid = True
                        if re.search('侧段', k):
                            right = True
                            mid = True
                        if re.search('下叶', k) or re.search('背段', k) or re.search('[外后]基底段', k):
                            down = True
                        if re.search('前基底段|(?<!前)内基底段', k):
                            right = True
                            down = True
                        if re.search('前内基底段', k):
                            left = True
                            down = True
                        if left:
                            if up:
                                col.append(15)
                            elif down:
                                col.append(16)
                            else:
                                col.append(14)
                        elif right:
                            if up:
                                col.append(18)
                            elif mid:
                                col.append(19)
                            elif down:
                                col.append(20)
                            else:
                                col.append(17)
                        elif double:
                            col.append(21)
                        left = False  # 标志位重置False
                        right = False
                        double = False
                        up = False
                        mid = False
                        down = False
                    print('col =', col)
                    # ------确定多发-------
                    if re.search('多发|[两多三四]枚|少许', sentence):
                        for k in col:
                            sheet.cell(row=i + 15, column=k).value = 'M'
                    # ------确定大小------
                    isSized = False  # 是否有大小信息
                    reg = '\d*\.?\d*(?:-|\*)?\d*\.?\d*(?:mm|cm)'
                    size_str = re.search(reg, sentence)
                    if not size_str:
                        if j != len(sentences) - 1:
                            size_str = re.search(reg, sentences[j + 1])
                            if not size_str:  # 没有大小信息填A
                                for k in col:
                                    if not sheet.cell(row=i + 15, column=k).value:
                                        sheet.cell(row=i + 15, column=k).value = 'A'
                            else:
                                isSized = True
                    else:
                        isSized = True
                    if isSized:
                        size_str = size_str.group()
                        if re.search('-', size_str):  # 有-则取-后面的数作为大小
                            result = re.search('(?<=-)\d*\.?\d*', size_str).group()
                            if re.search('cm', size_str):
                                if re.search('\.', result) and len(result.split('.')[1]) > 1:
                                    result = str(float(result) * 10)
                                else:
                                    result = str(int(float(result) * 10))
                        elif re.search('\*', size_str):
                            if re.search('mm', size_str):
                                result = re.search('[^mc]+', size_str).group()
                            else:  # cm
                                size_str = re.search('[^mc]+', size_str).group()
                                if re.search('\.', size_str.split('*')[0]) and len(
                                        size_str.split('*')[0].split('.')[1]) > 1:
                                    first = float(size_str.split('*')[0]) * 10
                                else:
                                    first = int(float(size_str.split('*')[0]) * 10)
                                if re.search('\.', size_str.split('*')[1]) and len(
                                        size_str.split('*')[1].split('.')[1]) > 1:
                                    second = float(size_str.split('*')[1]) * 10
                                else:
                                    second = int(float(size_str.split('*')[1]) * 10)
                                result = str(first) + '*' + str(second)
                        else:  # 单独数字
                            if re.search('mm', size_str):
                                result = re.search('[^mc]+', size_str).group()
                            else:
                                if re.search('\.', re.search('[^mc]+', size_str).group()) and len(
                                        re.search('[^mc]+', size_str).group().split('.')[1]) > 1:
                                    result = str(float(re.search('[^mc]+', size_str).group()) * 10)
                                else:
                                    result = str(int(float(re.search('[^mc]+', size_str).group()) * 10))
                        for k in col:
                            if sheet.cell(row=i + 15, column=k).value == 'M':
                                sheet.cell(row=i + 15, column=k).value += result
                            else:
                                sheet.cell(row=i + 15, column=k).value = result
                    col = []  # 清空位置列表

    workbook.save(path)
